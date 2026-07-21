from __future__ import annotations

from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from metadata import parse_column_guide_df, build_metadata_from_sav_info, merge_metadata
from utils import guess_rank_group_name, sort_var_names

QUESTION_SHEET = "문항설정"
BANNER_SHEET = "배너설정"
CHART_SHEET = "그래프설정"

TYPE_MAP = {
    "단일응답": "범주형",
    "복수응답": "다중응답",
    "4점척도": "척도형",
    "5점척도": "척도형",
    "7점척도": "척도형",
    "순위형": "순위형",
    "주관식": "개방형",
    "분석제외": "분석제외",
}


def _is_question_row(value) -> bool:
    text = str(value or "").strip()
    if not text or text.lower() == "nan":
        return False
    try:
        float(text)
        return False
    except Exception:
        return True


def load_survey_settings(uploaded_file, data_df: pd.DataFrame | None = None, sav_meta: dict | None = None) -> dict:
    if uploaded_file is None:
        raise ValueError("조사설정 엑셀파일이 없습니다.")
    uploaded_file.seek(0)
    xls = pd.ExcelFile(uploaded_file)
    missing_sheets = [s for s in (QUESTION_SHEET, BANNER_SHEET) if s not in xls.sheet_names]
    if missing_sheets:
        raise ValueError(f"조사설정 파일에 필요한 시트가 없습니다: {', '.join(missing_sheets)}")

    question_df = pd.read_excel(xls, sheet_name=QUESTION_SHEET, dtype=object).dropna(how="all")
    banner_df = pd.read_excel(xls, sheet_name=BANNER_SHEET, dtype=object).dropna(how="all")
    chart_df = pd.read_excel(xls, sheet_name=CHART_SHEET, dtype=object).dropna(how="all") if CHART_SHEET in xls.sheet_names else pd.DataFrame()

    required_q = ["QtnType", "문항/보기번호", "내용", "문항유형", "배너설정"]
    required_b = ["배너설정", "배너명", "기준변수", "출력순서"]
    missing_q = [c for c in required_q if c not in question_df.columns]
    missing_b = [c for c in required_b if c not in banner_df.columns]
    if missing_q or missing_b:
        raise ValueError("필수 열이 없습니다. " + "; ".join(filter(None, [f"문항설정: {missing_q}" if missing_q else "", f"배너설정: {missing_b}" if missing_b else ""])))

    if "문항영역" not in question_df.columns:
        question_df["문항영역"] = "미분류"
    question_df = question_df[required_q + ["문항영역"]].copy()
    question_rows = question_df[question_df["문항/보기번호"].map(_is_question_row)].copy()
    question_rows["var"] = question_rows["문항/보기번호"].astype(str).str.strip().str.lstrip("/")
    question_rows["문항유형"] = question_rows["문항유형"].fillna("").astype(str).str.strip()
    question_rows["배너설정"] = question_rows["배너설정"].fillna("전체값만 출력").astype(str).str.strip()
    question_rows["문항영역"] = question_rows["문항영역"].fillna("미분류").astype(str).str.strip().replace("", "미분류")

    invalid_types = sorted(set(question_rows.loc[~question_rows["문항유형"].isin(TYPE_MAP), "문항유형"]) - {""})
    if invalid_types:
        raise ValueError("지원하지 않는 문항유형이 있습니다: " + ", ".join(invalid_types))
    missing_type_vars = question_rows.loc[question_rows["문항유형"].eq(""), "var"].tolist()
    if missing_type_vars:
        raise ValueError("문항유형이 비어 있는 문항이 있습니다: " + ", ".join(missing_type_vars[:20]))

    guide_df = question_df[["QtnType", "문항/보기번호", "내용"]].copy()
    guide_metadata = parse_column_guide_df(guide_df)
    sav_metadata = build_metadata_from_sav_info(data_df, sav_meta) if data_df is not None else {}
    metadata = merge_metadata(guide_metadata, sav_metadata)

    question_order = question_rows["var"].tolist()
    type_map = {row["var"]: TYPE_MAP[row["문항유형"]] for _, row in question_rows.iterrows()}
    scale_vars = [v for v in question_order if type_map.get(v) == "척도형"]
    dep_vars = [v for v in question_order if type_map.get(v) in {"범주형", "척도형", "순위형"}]

    mr_group_map = {}
    for var in question_order:
        if type_map.get(var) == "다중응답":
            mr_group_map.setdefault(guess_rank_group_name(var), []).append(var)
    mr_group_map = {k: sort_var_names(v) for k, v in mr_group_map.items() if v}

    banner_df = banner_df.copy()
    banner_df["배너설정"] = banner_df["배너설정"].fillna("").astype(str).str.strip()
    banner_df["배너명"] = banner_df["배너명"].fillna("").astype(str).str.strip()
    banner_df["기준변수"] = banner_df["기준변수"].fillna("").astype(str).str.strip()
    banner_df["출력순서"] = pd.to_numeric(banner_df["출력순서"], errors="coerce").fillna(999999)
    banner_df = banner_df.sort_values(["배너설정", "출력순서"], kind="stable")

    banner_groups: dict[str, list[dict]] = {}
    for _, row in banner_df.iterrows():
        group = row["배너설정"]
        var = row["기준변수"]
        label = row["배너명"] or var
        if not group or not var:
            continue
        banner_groups.setdefault(group, []).append({"var": var, "label": label, "children": []})

    banner_by_question = {row["var"]: row["배너설정"] for _, row in question_rows.iterrows()}
    section_by_question = {row["var"]: row["문항영역"] for _, row in question_rows.iterrows()}
    for group_name, vars_ in mr_group_map.items():
        section_by_question[group_name] = next((section_by_question.get(v) for v in vars_ if section_by_question.get(v)), "미분류")

    graph_settings = {}
    if not chart_df.empty and "문항번호" in chart_df.columns:
        for _, row in chart_df.iterrows():
            qid = str(row.get("문항번호", "") or "").strip()
            if not qid:
                continue
            graph_settings[qid] = {
                "color": str(row.get("막대색상", "") or "").strip(),
                "linebreak": str(row.get("줄바꿈", "") or "").strip(),
            }
    profile_vars = []
    for nodes in banner_groups.values():
        for node in nodes:
            if node["var"] not in profile_vars:
                profile_vars.append(node["var"])

    errors = []
    warnings = []
    skipped_analysis_vars = []
    skipped_banner_vars = []
    if data_df is not None:
        data_cols = set(map(str, data_df.columns))
        configured = set(question_order)

        skipped_analysis_vars = sorted(
            v for v in configured
            if v not in data_cols and type_map.get(v) != "분석제외"
        )
        skipped_banner_vars = sorted(v for v in profile_vars if v not in data_cols)

        if skipped_analysis_vars:
            warnings.append(
                "데이터에 없는 분석 문항은 건너뜁니다: "
                + ", ".join(skipped_analysis_vars[:30])
                + (f" 외 {len(skipped_analysis_vars) - 30}개" if len(skipped_analysis_vars) > 30 else "")
            )
        if skipped_banner_vars:
            warnings.append(
                "데이터에 없는 배너변수는 건너뜁니다: "
                + ", ".join(skipped_banner_vars[:30])
                + (f" 외 {len(skipped_banner_vars) - 30}개" if len(skipped_banner_vars) > 30 else "")
            )

        # 실제 데이터에 존재하는 변수만 이후 분석 대상으로 사용한다.
        dep_vars = [v for v in dep_vars if v in data_cols]
        scale_vars = [v for v in scale_vars if v in data_cols]

        filtered_mr_group_map = {}
        for group_name, vars_ in mr_group_map.items():
            available_vars = [v for v in vars_ if v in data_cols]
            if available_vars:
                filtered_mr_group_map[group_name] = available_vars
        mr_group_map = filtered_mr_group_map

        filtered_banner_groups = {}
        for group_name, nodes in banner_groups.items():
            available_nodes = [node for node in nodes if node.get("var") in data_cols]
            if available_nodes:
                filtered_banner_groups[group_name] = available_nodes
        banner_groups = filtered_banner_groups
        profile_vars = [v for v in profile_vars if v in data_cols]

        unused_data = [c for c in data_df.columns if c not in configured]
        if unused_data:
            warnings.append(f"설정파일에 없는 데이터 컬럼 {len(unused_data)}개는 분석에서 제외됩니다.")

    return {
        "question_df": question_df,
        "question_rows": question_rows,
        "banner_df": banner_df,
        "metadata": metadata,
        "question_order": question_order,
        "type_map": type_map,
        "scale_vars": scale_vars,
        "dep_vars": dep_vars,
        "mr_group_map": mr_group_map,
        "selected_mr_groups": list(mr_group_map),
        "banner_groups": banner_groups,
        "banner_by_question": banner_by_question,
        "section_by_question": section_by_question,
        "graph_settings": graph_settings,
        "profile_vars": profile_vars,
        "errors": errors,
        "warnings": warnings,
        "skipped_analysis_vars": skipped_analysis_vars,
        "skipped_banner_vars": skipped_banner_vars,
    }


def build_settings_from_questionnaire(template_path: str | Path, guide_df: pd.DataFrame) -> bytes:
    wb = load_workbook(template_path)
    ws = wb[QUESTION_SHEET]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
        for cell in row:
            cell.value = None

    for idx, (_, row) in enumerate(guide_df.iterrows(), start=2):
        item = row.get("문항/보기번호", "")
        ws.cell(idx, 1, row.get("QtnType", ""))
        ws.cell(idx, 2, item)
        ws.cell(idx, 3, row.get("내용", ""))
        if _is_question_row(item):
            ws.cell(idx, 4, "단일응답")
            ws.cell(idx, 5, "전체값만 출력")
            ws.cell(idx, 6, "미분류")

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
