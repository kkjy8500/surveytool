from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
import pandas as pd
import re

from utils import clean_sheet_name, sort_var_names
from metadata import get_var_label


TABLE_GAP_ROWS = 2
OUTPUT_SHEET_NAME = "통계표"


def get_fill_by_name(color_name: str):
    color_map = {
        "없음": None,
        "연노랑": "FFF2CC",
        "연파랑": "D9EAF7",
        "연회색": "EDEDED",
        "연주황": "FCE4D6",
        "연초록": "E2F0D9",
    }
    hex_color = color_map.get(color_name)
    if not hex_color:
        return None
    return PatternFill("solid", fgColor=hex_color)


def is_profile_sheet_name(sheet_name: str) -> bool:
    return str(sheet_name).startswith("응답자특성")


def _resolve_table_title(table_name: str, metadata: dict) -> str:
    if table_name == "응답자특성":
        return "응답자 특성표"
    if table_name == "응답자특성_원본":
        return "응답자 특성표 | 원본"
    if table_name == "응답자특성_가중치":
        return "응답자 특성표 | 가중치 적용본"
    if table_name == "다문항요약":
        return "주요 문항 요약표"
    
    label = get_var_label(table_name, metadata)
    # Point 2: Improved Table Title (Q1. Label)
    if re.match(r"^[A-Za-z]+\d+(_\d+)?", table_name):
        return f"{table_name}. {label}"
    return f"{table_name} | {label}"


def _ordered_table_items(table_dict: dict) -> list[tuple[str, pd.DataFrame]]:
    """
    Consolidates and sorts tables in the correct order:
    1. 다문항요약 (Summary Table)
    2. 응답자특성 (Profile Tables)
    3. Other questions in natural sort order.
    """
    ordered_items = []
    keys = list(table_dict.keys())

    # 1. Summary Table
    if "다문항요약" in table_dict:
        ordered_items.append(("다문항요약", table_dict["다문항요약"]))
        if "다문항요약" in keys: keys.remove("다문항요약")

    # 2. Profile Tables
    profile_table_names = [name for name in keys if is_profile_sheet_name(name)]
    for name in sort_var_names(profile_table_names):
        ordered_items.append((name, table_dict[name]))
        if name in keys: keys.remove(name)

    # 3. Rest of the questions
    for name in sort_var_names(keys):
        ordered_items.append((name, table_dict[name]))

    return ordered_items


def _is_numeric_stat_header(header) -> bool:
    header_text = str(header)
    return (
        header in ["N", "N(명)", "%", "평균", "표준편차"]
        or header_text.endswith("_N")
        or header_text.endswith("_%")
    )


def _coerce_excel_cell_value(value, header):
    """
    통계표의 숫자 칸은 엑셀에서 빈칸이 아니라 0으로 표시되도록 정리한다.
    단, 구분/보기 같은 텍스트 칸의 빈값은 병합 처리를 위해 그대로 빈칸으로 둔다.
    """
    if pd.isna(value):
        return 0 if _is_numeric_stat_header(header) else ""
    if isinstance(value, str) and value.strip() == "" and _is_numeric_stat_header(header):
        return 0
    return value


def _merge_same_cells_by_columns_in_range(ws, col_indices, start_row: int, end_row: int):
    if end_row < start_row:
        return

    for col_idx in col_indices:
        block_start = start_row
        prev_val = ws.cell(row=start_row, column=col_idx).value

        for row_idx in range(start_row + 1, end_row + 1):
            current_val = ws.cell(row=row_idx, column=col_idx).value
            if current_val != prev_val:
                if row_idx - block_start > 1 and prev_val not in [None, ""]:
                    ws.merge_cells(
                        start_row=block_start,
                        start_column=col_idx,
                        end_row=row_idx - 1,
                        end_column=col_idx,
                    )
                    ws.cell(row=block_start, column=col_idx).alignment = Alignment(
                        vertical="center", horizontal="center"
                    )
                block_start = row_idx
                prev_val = current_val

        if end_row - block_start >= 1 and prev_val not in [None, ""]:
            ws.merge_cells(
                start_row=block_start,
                start_column=col_idx,
                end_row=end_row,
                end_column=col_idx,
            )
            ws.cell(row=block_start, column=col_idx).alignment = Alignment(
                vertical="center", horizontal="center"
            )


def _apply_table_style(
    ws,
    table_name: str,
    sheet_title: str,
    title_row: int,
    header_row: int,
    data_start_row: int,
    data_end_row: int,
    table_width: int,
    title_bold: bool,
    highlight_total: bool,
    stat_fill_name: str,
    decimals: int | None = None,
    pct_decimals: int | None = 1,
    stat_decimals: int | None = 2,
    n_in_parentheses: bool = False,
):
    # Point 3: Header Style (Light Blue Background, Navy Font)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True, color="000080")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Point 5: Total Row Style (Light Gray)
    total_fill = PatternFill("solid", fgColor="EDEDED")
    total_font = Font(bold=True)
    
    # Point 4: Banner Column Style (Very Light Gray)
    banner_fill = PatternFill("solid", fgColor="F2F2F2")
    banner_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    stat_fill = get_fill_by_name(stat_fill_name)
    thin_side = Side(style="thin", color="D9D9D9")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # Title Style (Bold, Large)
    ws.cell(row=title_row, column=1, value=sheet_title)
    if table_width > 1:
        ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=table_width)
    ws.cell(row=title_row, column=1).font = Font(bold=True, size=13)
    ws.cell(row=title_row, column=1).alignment = Alignment(horizontal="left", vertical="center")

    # Apply Header Styles
    for col_idx in range(1, table_width + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border

    # Apply Base Styles to Data Rows
    for row_idx in range(data_start_row, data_end_row + 1):
        is_total_row = str(ws.cell(row=row_idx, column=1).value).strip() == "전체"
        
        for col_idx in range(1, table_width + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            header = str(ws.cell(row=header_row, column=col_idx).value)
            
            # Highlight Total Row
            if is_total_row and highlight_total:
                cell.fill = total_fill
                cell.font = total_font
            
            # Point 4: Highlight Banner/Group Columns
            elif header.startswith("구분") or header in ["변수", "문항", "구분1", "구분2", "구분3"]:
                cell.fill = banner_fill
                cell.alignment = banner_align
                if header == "구분1" or header == "문항":
                    cell.font = Font(bold=True)

    # Point 6 & 7: Number Formats
    pct_format = "0" if pct_decimals == 0 else "0." + ("0" * pct_decimals)
    stat_format = "0" if stat_decimals == 0 else "0." + ("0" * stat_decimals)
    effective_n_in_parentheses = False if is_profile_sheet_name(table_name) else n_in_parentheses
    n_format = r"\(#,##0\)" if effective_n_in_parentheses else "#,##0"

    header_map = {col_idx: ws.cell(row=header_row, column=col_idx).value for col_idx in range(1, table_width + 1)}

    for row in ws.iter_rows(min_row=data_start_row, max_row=data_end_row, min_col=1, max_col=table_width):
        for cell in row:
            header = str(header_map.get(cell.column))
            if not isinstance(cell.value, (int, float)):
                continue

            if header in ["N", "N(명)"] or header.endswith("_N"):
                cell.number_format = n_format
            elif header == "%" or header.endswith("_%"):
                cell.number_format = pct_format
            elif header in ["평균", "표준편차", "100점 환산"] or "평균" in header:
                cell.number_format = stat_format
                if stat_fill:
                    cell.fill = stat_fill

    # Merging Group Columns
    group_col_indices = []
    for col_idx in range(1, table_width + 1):
        header = str(ws.cell(row=header_row, column=col_idx).value)
        if header.startswith("구분"):
            group_col_indices.append(col_idx)

    if group_col_indices:
        _merge_same_cells_by_columns_in_range(ws, group_col_indices, start_row=data_start_row, end_row=data_end_row)


def _find_total_row_in_range(ws, start_row: int, end_row: int):
    for row_idx in range(start_row, end_row + 1):
        if ws.cell(row=row_idx, column=1).value == "전체":
            return row_idx
    return None


def add_excel_chart_to_sheet(
    ws,
    dep_var: str,
    metadata: dict,
    header_row: int = 2,
    data_start_row: int = 3,
    data_end_row: int | None = None,
    table_width: int | None = None,
    title_row: int = 1,
):
    if data_end_row is None:
        data_end_row = ws.max_row
    if table_width is None:
        table_width = ws.max_column

    header_map = {idx: ws.cell(row=header_row, column=idx).value for idx in range(1, table_width + 1)}
    total_row = _find_total_row_in_range(ws, data_start_row, data_end_row)
    if total_row is None:
        return False

    pct_cols = []
    cat_labels = []
    for idx, header in header_map.items():
        if isinstance(header, str) and header.endswith("_%"):
            pct_cols.append(idx)
            cat_labels.append(header[:-2])

    if len(pct_cols) < 2:
        return False

    chart_data_start_col = table_width + 2
    chart_label_col = chart_data_start_col
    chart_value_col = chart_data_start_col + 1

    ws.cell(row=header_row, column=chart_label_col, value="차트라벨")
    ws.cell(row=header_row, column=chart_value_col, value="차트값")

    for offset, (label, pct_col) in enumerate(zip(cat_labels, pct_cols), start=1):
        row_idx = header_row + offset
        ws.cell(row=row_idx, column=chart_label_col, value=label)
        ws.cell(row=row_idx, column=chart_value_col, value=ws.cell(row=total_row, column=pct_col).value)

    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.y_axis.title = ""
    chart.x_axis.title = "%"
    chart.height = max(6, min(14, len(cat_labels) * 0.6 + 2))
    chart.width = 11
    chart.title = f"{dep_var} | {get_var_label(dep_var, metadata)}"
    chart.legend = None
    chart.varyColors = True

    data = Reference(ws, min_col=chart_value_col, min_row=header_row, max_row=header_row + len(cat_labels))
    cats = Reference(ws, min_col=chart_label_col, min_row=header_row + 1, max_row=header_row + len(cat_labels))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.dLbls = DataLabelList()
    chart.dLbls.showVal = True
    chart.dLbls.showPercent = False

    anchor_col = get_column_letter(chart_data_start_col + 3)
    ws.add_chart(chart, f"{anchor_col}{title_row}")
    return True


def _write_table_block(ws, table_name: str, table_df: pd.DataFrame, metadata: dict, start_row: int, **style_kwargs) -> int:
    if table_df is None:
        table_df = pd.DataFrame()
    table_df = table_df.copy()

    headers = [str(col) for col in table_df.columns]
    table_width = max(1, len(headers))
    title_row = start_row
    header_row = start_row + 1
    data_start_row = start_row + 2

    sheet_title = _resolve_table_title(table_name, metadata)

    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col_idx, value=header)

    for row_offset, (_, row) in enumerate(table_df.iterrows(), start=0):
        row_idx = data_start_row + row_offset
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col_idx, value=_coerce_excel_cell_value(row.get(header), header))

    data_end_row = max(data_start_row, data_start_row + len(table_df) - 1)

    _apply_table_style(
        ws=ws,
        table_name=table_name,
        sheet_title=sheet_title,
        title_row=title_row,
        header_row=header_row,
        data_start_row=data_start_row,
        data_end_row=data_end_row,
        table_width=table_width,
        **style_kwargs,
    )

    return data_end_row


def _apply_column_widths(ws):
    """Point 9: Automated column width optimization."""
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        header_val = str(ws.cell(row=2, column=col_idx).value or "") # Assume header is at row 2
        
        # Categorical columns get more width
        if header_val.startswith("구분") or header_val in ["문항", "변수"]:
            ws.column_dimensions[col_letter].width = 32
        # Numeric columns are narrow
        elif _is_numeric_stat_header(header_val) or "평균" in header_val:
            ws.column_dimensions[col_letter].width = 12
        else:
            ws.column_dimensions[col_letter].width = 15


def export_to_excel(
    table_dict: dict,
    metadata: dict,
    title_bold: bool,
    highlight_total: bool,
    stat_fill_name: str,
    decimals: int | None = None,
    pct_decimals: int | None = 1,
    stat_decimals: int | None = 2,
    add_excel_chart: bool = False,
    n_in_parentheses: bool = False,
    audit_trail: dict | None = None,
):
    """
    Main export engine with Index and Audit Trail support (Point 11, 12).
    """
    output = BytesIO()
    workbook = Workbook()
    
    # --- 1. Index Sheet ---
    index_ws = workbook.active
    index_ws.title = "목차"
    index_ws.cell(row=1, column=1, value="통계표 목차").font = Font(bold=True, size=14)
    index_ws.cell(row=3, column=1, value="번호").font = Font(bold=True)
    index_ws.cell(row=3, column=2, value="문항ID").font = Font(bold=True)
    index_ws.cell(row=3, column=3, value="문항명").font = Font(bold=True)
    
    # --- 2. Main Stats Sheet ---
    ws = workbook.create_sheet(clean_sheet_name(OUTPUT_SHEET_NAME))
    ordered_items = _ordered_table_items(table_dict or {})
    
    current_row = 2 # Start from row 2 for better spacing
    
    style_kwargs = {
        "title_bold": title_bold,
        "highlight_total": highlight_total,
        "stat_fill_name": stat_fill_name,
        "decimals": decimals,
        "pct_decimals": pct_decimals,
        "stat_decimals": stat_decimals,
        "n_in_parentheses": n_in_parentheses,
    }

    for idx, (table_name, table_df) in enumerate(ordered_items, start=1):
        table_df = table_df.copy() if table_df is not None else pd.DataFrame()
        
        title_row = current_row
        
        # Add entry to Index sheet
        index_row = idx + 3
        index_ws.cell(row=index_row, column=1, value=idx)
        index_ws.cell(row=index_row, column=2, value=table_name)
        full_title = _resolve_table_title(table_name, metadata)
        index_ws.cell(row=index_row, column=3, value=full_title)
        
        # Hyperlink to the table
        link_target = f"'{ws.title}'!A{title_row}"
        index_ws.cell(row=index_row, column=3).hyperlink = f"#{link_target}"
        index_ws.cell(row=index_row, column=3).font = Font(color="0000FF", underline="single")

        data_end_row = _write_table_block(
            ws=ws,
            table_name=table_name,
            table_df=table_df,
            metadata=metadata,
            start_row=current_row,
            **style_kwargs,
        )

        if add_excel_chart and not is_profile_sheet_name(table_name):
            add_excel_chart_to_sheet(
                ws=ws,
                dep_var=table_name,
                metadata=metadata,
                header_row=header_row,
                data_start_row=data_start_row,
                data_end_row=data_end_row,
                table_width=table_width,
                title_row=title_row,
            )

        current_row = data_end_row + TABLE_GAP_ROWS + 1

    _apply_column_widths(ws)
    workbook.save(output)
    output.seek(0)
    return output
